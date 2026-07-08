import json
import io
from datetime import datetime, timedelta
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.core.files.base import ContentFile
from django.contrib.auth import get_user_model
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv

from .models import Report

User = get_user_model()


class ReportGenerator:
    """
    مولد التقارير المتكامل
    """
    
    def __init__(self, report_id):
        self.report = Report.objects.get(id=report_id)
    
    def generate(self):
        """توليد التقرير حسب النوع"""
        self.report.mark_processing()
        
        try:
            # استخراج البيانات حسب النوع
            data = self._extract_data()
            
            # توليد الملخص
            summary = self._generate_summary(data)
            
            # توليد الملف
            file = self._generate_file(data)
            
            # حفظ النتائج
            self.report.mark_completed(
                data=data,
                summary=summary,
                file=file
            )
            
            return True
            
        except Exception as e:
            self.report.mark_failed(str(e))
            raise
    
    def _extract_data(self):
        """استخراج البيانات حسب نوع التقرير"""
        data = {}
        
        if self.report.report_type == Report.Type.BOOKING:
            data = self._extract_booking_data()
        elif self.report.report_type == Report.Type.FINANCIAL:
            data = self._extract_financial_data()
        elif self.report.report_type == Report.Type.CUSTOMER:
            data = self._extract_customer_data()
        elif self.report.report_type == Report.Type.PHOTOGRAPHER:
            data = self._extract_photographer_data()
        elif self.report.report_type == Report.Type.PAYMENT:
            data = self._extract_payment_data()
        elif self.report.report_type == Report.Type.EXPENSE:
            data = self._extract_expense_data()
        elif self.report.report_type == Report.Type.PROFIT:
            data = self._extract_profit_data()
        elif self.report.report_type == Report.Type.DAILY:
            data = self._extract_daily_data()
        elif self.report.report_type == Report.Type.WEEKLY:
            data = self._extract_weekly_data()
        elif self.report.report_type == Report.Type.MONTHLY:
            data = self._extract_monthly_data()
        elif self.report.report_type == Report.Type.YEARLY:
            data = self._extract_yearly_data()
        
        return data
    
    def _extract_booking_data(self):
        """استخراج بيانات الحجوزات"""
        from bookings.models import Booking
        
        queryset = Booking.objects.select_related('customer', 'photographer')
        
        # تطبيق الفلاتر
        filters = self.report.filters
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        if filters.get('photographer_id'):
            queryset = queryset.filter(photographer_id=filters['photographer_id'])
        if filters.get('customer_id'):
            queryset = queryset.filter(customer_id=filters['customer_id'])
        if self.report.start_date:
            queryset = queryset.filter(date__gte=self.report.start_date.date())
        if self.report.end_date:
            queryset = queryset.filter(date__lte=self.report.end_date.date())
        
        # إحصائيات
        stats = {
            'total': queryset.count(),
            'confirmed': queryset.filter(status='confirmed').count(),
            'pending': queryset.filter(status='pending').count(),
            'cancelled': queryset.filter(status='cancelled').count(),
            'done': queryset.filter(status='done').count(),
            'total_revenue': queryset.filter(status='done').aggregate(Sum('price'))['price__sum'] or 0,
        }
        
        # التفاصيل
        details = []
        for booking in queryset:
            details.append({
                'id': booking.id,
                'customer': str(booking.customer),
                'photographer': str(booking.photographer),
                'date': booking.date.strftime('%Y-%m-%d'),
                'start_time': booking.start_time.strftime('%H:%M'),
                'end_time': booking.end_time.strftime('%H:%M'),
                'status': booking.get_status_display(),
                'price': float(booking.price),
            })
        
        return {
            'stats': stats,
            'details': details,
            'chart_data': {
                'labels': ['مؤكد', 'قيد الانتظار', 'ملغي', 'منتهي'],
                'values': [stats['confirmed'], stats['pending'], stats['cancelled'], stats['done']],
            }
        }
    
    def _extract_financial_data(self):
        """استخراج البيانات المالية"""
        from finance.models import Income, Expense
        
        filters = self.report.filters
        
        # الإيرادات
        income_queryset = Income.objects.all()
        if self.report.start_date:
            income_queryset = income_queryset.filter(income_date__gte=self.report.start_date)
        if self.report.end_date:
            income_queryset = income_queryset.filter(income_date__lte=self.report.end_date)
        
        # المصروفات
        expense_queryset = Expense.objects.all()
        if self.report.start_date:
            expense_queryset = expense_queryset.filter(expense_date__gte=self.report.start_date)
        if self.report.end_date:
            expense_queryset = expense_queryset.filter(expense_date__lte=self.report.end_date)
        
        total_income = income_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        total_expense = expense_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        
        # حسب المصدر
        income_by_source = dict(
            income_queryset.values('source').annotate(
                total=Sum('amount')
            ).values_list('source', 'total')
        )
        
        # حسب التصنيف
        expense_by_category = dict(
            expense_queryset.values('category').annotate(
                total=Sum('amount')
            ).values_list('category', 'total')
        )
        
        return {
            'stats': {
                'total_income': total_income,
                'total_expense': total_expense,
                'profit': total_income - total_expense,
                'income_count': income_queryset.count(),
                'expense_count': expense_queryset.count(),
            },
            'income_by_source': income_by_source,
            'expense_by_category': expense_by_category,
            'chart_data': {
                'income_labels': list(income_by_source.keys()),
                'income_values': list(income_by_source.values()),
                'expense_labels': list(expense_by_category.keys()),
                'expense_values': list(expense_by_category.values()),
            }
        }
    
    def _extract_customer_data(self):
        """استخراج بيانات العملاء"""
        from customers.models import Customer
        
        queryset = Customer.objects.all()
        
        if self.report.start_date:
            queryset = queryset.filter(created_at__gte=self.report.start_date)
        if self.report.end_date:
            queryset = queryset.filter(created_at__lte=self.report.end_date)
        
        total = queryset.count()
        vip = queryset.filter(customer_type=Customer.TYPE_VIP).count()
        
        details = []
        for customer in queryset:
            details.append({
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone,
                'email': customer.email or '',
                'type': customer.get_customer_type_display(),
                'bookings': customer.total_bookings,
                'spent': float(customer.total_spent),
                'created_at': customer.created_at.strftime('%Y-%m-%d'),
            })
        
        return {
            'stats': {
                'total': total,
                'vip': vip,
                'active': queryset.filter(is_active=True).count(),
                'new_this_period': queryset.count(),
            },
            'details': details,
        }
    
    def _extract_photographer_data(self):
        """استخراج بيانات المصورين"""
        from users.models import User
        
        photographers = User.objects.filter(role='photographer', is_active=True)
        
        details = []
        total_bookings = 0
        
        for photographer in photographers:
            bookings = photographer.bookings.all()
            if self.report.start_date:
                bookings = bookings.filter(date__gte=self.report.start_date.date())
            if self.report.end_date:
                bookings = bookings.filter(date__lte=self.report.end_date.date())
            
            count = bookings.count()
            total_bookings += count
            revenue = bookings.filter(status='done').aggregate(Sum('price'))['price__sum'] or 0
            
            details.append({
                'id': photographer.id,
                'name': photographer.get_full_name() or photographer.username,
                'bookings': count,
                'revenue': float(revenue),
                'rating': 0,  # يمكن إضافة نظام تقييم
            })
        
        return {
            'stats': {
                'total_photographers': photographers.count(),
                'total_bookings': total_bookings,
                'average_bookings': total_bookings / photographers.count() if photographers.exists() else 0,
            },
            'details': details,
        }
    
    def _extract_payment_data(self):
        """استخراج بيانات المدفوعات"""
        from payments.models import Payment
        
        queryset = Payment.objects.select_related('booking', 'booking__customer')
        
        if self.report.start_date:
            queryset = queryset.filter(payment_date__gte=self.report.start_date)
        if self.report.end_date:
            queryset = queryset.filter(payment_date__lte=self.report.end_date)
        
        stats = {
            'total': queryset.count(),
            'total_amount': queryset.aggregate(Sum('amount'))['amount__sum'] or 0,
            'paid': queryset.filter(status='paid').count(),
            'pending': queryset.filter(status='pending').count(),
            'failed': queryset.filter(status='failed').count(),
            'refunded': queryset.filter(status='refunded').count(),
        }
        
        details = []
        for payment in queryset:
            details.append({
                'id': payment.id,
                'booking_id': payment.booking.id if payment.booking else None,
                'customer': str(payment.booking.customer) if payment.booking and payment.booking.customer else '',
                'amount': float(payment.amount),
                'type': payment.get_payment_type_display(),
                'status': payment.get_status_display(),
                'method': payment.get_payment_method_display(),
                'date': payment.payment_date.strftime('%Y-%m-%d %H:%M'),
            })
        
        return {
            'stats': stats,
            'details': details,
        }
    
    def _extract_daily_data(self):
        """بيانات يومية"""
        today = timezone.now().date()
        return self._extract_data_for_date(today, today)
    
    def _extract_weekly_data(self):
        """بيانات أسبوعية"""
        today = timezone.now().date()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return self._extract_data_for_date(start, end)
    
    def _extract_monthly_data(self):
        """بيانات شهرية"""
        today = timezone.now().date()
        start = today.replace(day=1)
        if today.month == 12:
            end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        return self._extract_data_for_date(start, end)
    
    def _extract_yearly_data(self):
        """بيانات سنوية"""
        today = timezone.now().date()
        start = today.replace(month=1, day=1)
        end = today.replace(month=12, day=31)
        return self._extract_data_for_date(start, end)
    
    def _extract_data_for_date(self, start_date, end_date):
        """استخراج البيانات لنطاق زمني محدد"""
        from bookings.models import Booking
        from finance.models import Income, Expense
        
        # حجوزات
        bookings = Booking.objects.filter(date__range=[start_date, end_date])
        
        # إيرادات ومصروفات
        income = Income.objects.filter(income_date__range=[start_date, end_date])
        expense = Expense.objects.filter(expense_date__range=[start_date, end_date])
        
        total_income = income.aggregate(Sum('amount'))['amount__sum'] or 0
        total_expense = expense.aggregate(Sum('amount'))['amount__sum'] or 0
        
        return {
            'period': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d'),
            },
            'stats': {
                'bookings': bookings.count(),
                'confirmed': bookings.filter(status='confirmed').count(),
                'done': bookings.filter(status='done').count(),
                'cancelled': bookings.filter(status='cancelled').count(),
                'income': total_income,
                'expenses': total_expense,
                'profit': total_income - total_expense,
            }
        }
    
    def _extract_profit_data(self):
        """بيانات الأرباح"""
        from finance.models import Income, Expense
        
        filters = self.report.filters
        
        # الفترة
        start_date = self.report.start_date
        end_date = self.report.end_date
        
        if not start_date or not end_date:
            # آخر 30 يوم
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
        
        # البيانات اليومية
        daily_data = []
        current = start_date
        while current <= end_date:
            day_start = current
            day_end = current + timedelta(days=1)
            
            daily_income = Income.objects.filter(
                income_date__range=[day_start, day_end]
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            daily_expense = Expense.objects.filter(
                expense_date__range=[day_start, day_end]
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            daily_data.append({
                'date': current.strftime('%Y-%m-%d'),
                'income': daily_income,
                'expenses': daily_expense,
                'profit': daily_income - daily_expense,
            })
            
            current += timedelta(days=1)
        
        # إجمالي
        total_income = sum(d['income'] for d in daily_data)
        total_expense = sum(d['expenses'] for d in daily_data)
        
        return {
            'period': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d'),
                'days': (end_date - start_date).days + 1,
            },
            'summary': {
                'total_income': total_income,
                'total_expenses': total_expense,
                'total_profit': total_income - total_expense,
                'average_daily_profit': (total_income - total_expense) / ((end_date - start_date).days + 1),
            },
            'daily_data': daily_data,
            'chart_data': {
                'labels': [d['date'] for d in daily_data],
                'income': [d['income'] for d in daily_data],
                'expenses': [d['expenses'] for d in daily_data],
                'profit': [d['profit'] for d in daily_data],
            }
        }
    
    def _generate_summary(self, data):
        """توليد ملخص التقرير"""
        summary = {}
        
        if 'stats' in data:
            summary.update(data['stats'])
        
        if 'period' in data:
            summary['period'] = data['period']
        
        if 'total' in data.get('stats', {}):
            summary['total_items'] = data['stats']['total']
        
        return summary
    
    def _generate_file(self, data):
        """توليد ملف التقرير بالصيغة المطلوبة"""
        if self.report.format == Report.Format.PDF:
            return self._generate_pdf(data)
        elif self.report.format == Report.Format.EXCEL:
            return self._generate_excel(data)
        elif self.report.format == Report.Format.CSV:
            return self._generate_csv(data)
        elif self.report.format == Report.Format.JSON:
            return self._generate_json(data)
        else:
            return self._generate_pdf(data)
    
    def _generate_pdf(self, data):
        """توليد ملف PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # عنوان التقرير
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        story.append(Paragraph(self.report.title, title_style))
        
        # معلومات التقرير
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666'),
        )
        story.append(Paragraph(f"النوع: {self.report.get_report_type_display()}", info_style))
        story.append(Paragraph(f"التاريخ: {timezone.now().strftime('%Y-%m-%d %H:%M')}", info_style))
        story.append(Spacer(1, 20))
        
        # الملخص
        if 'summary' in data:
            story.append(Paragraph("الملخص", styles['Heading2']))
            for key, value in data['summary'].items():
                if isinstance(value, (int, float)):
                    story.append(Paragraph(f"{key}: {value:,}", styles['Normal']))
                elif isinstance(value, str):
                    story.append(Paragraph(f"{key}: {value}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # الجدول
        if 'details' in data and data['details']:
            story.append(Paragraph("التفاصيل", styles['Heading2']))
            
            # استخراج العناوين
            headers = list(data['details'][0].keys())
            data_rows = [[h for h in headers]]
            
            for item in data['details']:
                row = []
                for h in headers:
                    val = item.get(h, '')
                    if isinstance(val, float):
                        val = f"{val:,.2f}"
                    row.append(str(val))
                data_rows.append(row)
            
            # إنشاء الجدول
            table = Table(data_rows)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ddd')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(table)
        
        doc.build(story)
        
        buffer.seek(0)
        filename = f"{self.report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return ContentFile(buffer.read(), name=filename)
    
    def _generate_excel(self, data):
        """توليد ملف Excel"""
        wb = Workbook()
        
        # الصفحة الأولى: الملخص
        ws1 = wb.active
        ws1.title = "الملخص"
        
        # عنوان
        ws1['A1'] = self.report.title
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        # معلومات
        ws1['A3'] = f"النوع: {self.report.get_report_type_display()}"
        ws1['A4'] = f"التاريخ: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        # الملخص
        row = 6
        if 'summary' in data:
            ws1['A6'] = "الملخص"
            ws1['A6'].font = Font(bold=True, size=14)
            row = 7
            for key, value in data['summary'].items():
                ws1[f'A{row}'] = str(key)
                ws1[f'B{row}'] = str(value)
                row += 1
        
        # الصفحة الثانية: التفاصيل
        if 'details' in data and data['details']:
            ws2 = wb.create_sheet("التفاصيل")
            
            # العناوين
            headers = list(data['details'][0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # البيانات
            for row, item in enumerate(data['details'], 2):
                for col, header in enumerate(headers, 1):
                    val = item.get(header, '')
                    if isinstance(val, float):
                        val = round(val, 2)
                    ws2.cell(row=row, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"{self.report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return ContentFile(buffer.read(), name=filename)
    
    def _generate_csv(self, data):
        """توليد ملف CSV"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # كتابة المعلومات
        writer.writerow([self.report.title])
        writer.writerow([f"النوع: {self.report.get_report_type_display()}"])
        writer.writerow([f"التاريخ: {timezone.now().strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow([])
        
        # الملخص
        if 'summary' in data:
            writer.writerow(["الملخص"])
            for key, value in data['summary'].items():
                writer.writerow([key, value])
            writer.writerow([])
        
        # التفاصيل
        if 'details' in data and data['details']:
            headers = list(data['details'][0].keys())
            writer.writerow(headers)
            
            for item in data['details']:
                row = []
                for h in headers:
                    val = item.get(h, '')
                    if isinstance(val, float):
                        val = round(val, 2)
                    row.append(val)
                writer.writerow(row)
        
        buffer.seek(0)
        filename = f"{self.report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return ContentFile(buffer.getvalue().encode('utf-8'), name=filename)
    
    def _generate_json(self, data):
        """توليد ملف JSON"""
        import json
        
        content = json.dumps(data, ensure_ascii=False, indent=2)
        filename = f"{self.report.title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        return ContentFile(content.encode('utf-8'), name=filename)